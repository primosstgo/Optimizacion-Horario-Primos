from pulp import LpMaximize, LpProblem, LpStatus, lpSum, LpVariable
import random
import openpyxl
import re
from Disponibilidad_primos import Horarios_primos

#defincion del modelo
model = LpProblem(name="Horario", sense=LpMaximize)

#definicion de variables
dias = ["L","M","X","J","V"]
Primos = ["P1","P2","P3","P4","P5","P6","P7","P8","P9","P10","P11","P12","P13","P14","P15","P16"]

for x in range(0,16):
    for y in range(0,5):
        for z in range(1,8):
            nombre_variable = f"{Primos[x]}{dias[y]}{z}"
            globals()[nombre_variable] = Horarios_primos[x][1][y][z-1]

# Cantidad de horas por Primo
cantidad_de_primos = 16 # solo funciona con 16 (de momento)

P = LpVariable.dicts("P", range(1, cantidad_de_primos+1), lowBound=4, upBound=5, cat="Integer")

#Valor de cada dia

L = [LpVariable(name=f"L{i}", lowBound=0, cat="Integer") for i in range(1, 8)]
M = [LpVariable(name=f"M{i}", lowBound=0, cat="Integer") for i in range(1, 8)]
X = [LpVariable(name=f"X{i}", lowBound=0, cat="Integer") for i in range(1, 8)]
J = [LpVariable(name=f"J{i}", lowBound=0, cat="Integer") for i in range(1, 8)]
V = [LpVariable(name=f"V{i}", lowBound=0, cat="Integer") for i in range(1, 8)]
J = [LpVariable(name=f"J{i}", lowBound=0, cat="Integer") for i in range(1, 17)]

#turno elegido por P[Primo]_[Dia][bloque]

for i in range(1,17):
    for x in dias:
        for z in range(1,8):
            globals()["P"+str(i)+"_"+str(x)+str(z)] = LpVariable(name="P"+str(i)+"_"+str(x)+str(z), cat="Binary")


##condiciones

##funcion obj
model += L[0] + L[1] + L[2] + L[3] + L[4] + L[5] + L[6]  + M[0] +  M[1] + M[2] + M[3] + M[4] + M[5] + M[6] + X[0] + X[1] + X[2] + X[3] + X[4] + X[5] + X[6] + J[0] + J[1] + J[2] + J[3] + J[4] + J[5] + J[6] + V[0] + V[1] + V[2] + V[3] + V[4] + V[5] + V[6] 
## valor del dia
model += (L[0] == P1L1*P1_L1 + P2L1*P2_L1 + P3L1*P3_L1 + P4L1*P4_L1 + P5L1*P5_L1 + P6L1*P6_L1 + P7L1*P7_L1 + P8L1*P8_L1 + P9L1*P9_L1 + P10L1*P10_L1 + P11L1*P11_L1 +  P12L1*P2_L1 + P13L1*P13_L1 + P14L1*P14_L1 + P15L1*P15_L1 + P16L1*P16_L1 , "SL1")
model += (L[1] == P1L2*P1_L2 + P2L2*P2_L2 + P3L2*P3_L2 + P4L2*P4_L2 + P5L2*P5_L2 + P6L2*P6_L2 + P7L2*P7_L2 + P8L2*P8_L2 + P9L2*P9_L2 + P10L2*P10_L2 + P11L2*P11_L2 +  P12L2*P2_L2 + P13L2*P13_L2 + P14L2*P14_L2 + P15L2*P15_L2 + P16L2*P16_L2 , "SL2")
model += (L[2] == P1L3*P1_L3 + P2L3*P2_L3 + P3L3*P3_L3 + P4L3*P4_L3 + P5L3*P5_L3 + P6L3*P6_L3 + P7L3*P7_L3 + P8L3*P8_L3 + P9L3*P9_L3 + P10L3*P10_L3 + P11L3*P11_L3 +  P12L3*P2_L3 + P13L3*P13_L3 + P14L3*P14_L3 + P15L3*P15_L3 + P16L3*P16_L3 , "SL3")
model += (L[3] == P1L4*P1_L4 + P2L4*P2_L4 + P3L4*P3_L4 + P4L4*P4_L4 + P5L4*P5_L4 + P6L4*P6_L4 + P7L4*P7_L4 + P8L4*P8_L4 + P9L4*P9_L4 + P10L4*P10_L4 + P11L4*P11_L4 +  P12L4*P2_L4 + P13L4*P13_L4 + P14L4*P14_L4 + P15L4*P15_L4 + P16L4*P16_L4 , "SL4")
model += (L[4] == P1L5*P1_L5 + P2L5*P2_L5 + P3L5*P3_L5 + P4L5*P4_L5 + P5L5*P5_L5 + P6L5*P6_L5 + P7L5*P7_L5 + P8L5*P8_L5 + P9L5*P9_L5 + P10L5*P10_L5 + P11L5*P11_L5 +  P12L5*P2_L5 + P13L5*P13_L5 + P14L5*P14_L5 + P15L5*P15_L5 + P16L5*P16_L5 , "SL5")
model += (L[5] == P1L6*P1_L6 + P2L6*P2_L6 + P3L6*P3_L6 + P4L6*P4_L6 + P5L6*P5_L6 + P6L6*P6_L6 + P7L6*P7_L6 + P8L6*P8_L6 + P9L6*P9_L6 + P10L6*P10_L6 + P11L6*P11_L6 +  P12L6*P2_L6 + P13L6*P13_L6 + P14L6*P14_L6 + P15L6*P15_L6 + P16L6*P16_L6 , "SL6")
model += (L[6] == P1L7*P1_L7 + P2L7*P2_L7 + P3L7*P3_L7 + P4L7*P4_L7 + P5L7*P5_L7 + P6L7*P6_L7 + P7L7*P7_L7 + P8L7*P8_L7 + P9L7*P9_L7 + P10L7*P10_L7 + P11L7*P11_L7 +  P12L7*P2_L7 + P13L7*P13_L7 + P14L7*P14_L7 + P15L7*P15_L7 + P16L7*P16_L7 , "SL7")

model += (M[0] == P1M1*P1_M1 + P2M1*P2_M1 + P3M1*P3_M1 + P4M1*P4_M1 + P5M1*P5_M1 + P6M1*P6_M1 + P7M1*P7_M1 + P8M1*P8_M1 + P9M1*P9_M1 + P10M1*P10_M1 + P11M1*P11_M1 +  P12M1*P2_M1 + P13M1*P13_M1 + P14M1*P14_M1 + P15M1*P15_M1 + P16M1*P16_M1 , "SM1")
model += (M[1] == P1M2*P1_M2 + P2M2*P2_M2 + P3M2*P3_M2 + P4M2*P4_M2 + P5M2*P5_M2 + P6M2*P6_M2 + P7M2*P7_M2 + P8M2*P8_M2 + P9M2*P9_M2 + P10M2*P10_M2 + P11M2*P11_M2 +  P12M2*P2_M2 + P13M2*P13_M2 + P14M2*P14_M2 + P15M2*P15_M2 + P16M2*P16_M2 , "SM2")
model += (M[2] == P1M3*P1_M3 + P2M3*P2_M3 + P3M3*P3_M3 + P4M3*P4_M3 + P5M3*P5_M3 + P6M3*P6_M3 + P7M3*P7_M3 + P8M3*P8_M3 + P9M3*P9_M3 + P10M3*P10_M3 + P11M3*P11_M3 +  P12M3*P2_M3 + P13M3*P13_M3 + P14M3*P14_M3 + P15M3*P15_M3 + P16M3*P16_M3 , "SM3")
model += (M[3] == P1M4*P1_M4 + P2M4*P2_M4 + P3M4*P3_M4 + P4M4*P4_M4 + P5M4*P5_M4 + P6M4*P6_M4 + P7M4*P7_M4 + P8M4*P8_M4 + P9M4*P9_M4 + P10M4*P10_M4 + P11M4*P11_M4 +  P12M4*P2_M4 + P13M4*P13_M4 + P14M4*P14_M4 + P15M4*P15_M4 + P16M4*P16_M4 , "SM4")
model += (M[4] == P1M5*P1_M5 + P2M5*P2_M5 + P3M5*P3_M5 + P4M5*P4_M5 + P5M5*P5_M5 + P6M5*P6_M5 + P7M5*P7_M5 + P8M5*P8_M5 + P9M5*P9_M5 + P10M5*P10_M5 + P11M5*P11_M5 +  P12M5*P2_M5 + P13M5*P13_M5 + P14M5*P14_M5 + P15M5*P15_M5 + P16M5*P16_M5 , "SM5")
model += (M[5] == P1M6*P1_M6 + P2M6*P2_M6 + P3M6*P3_M6 + P4M6*P4_M6 + P5M6*P5_M6 + P6M6*P6_M6 + P7M6*P7_M6 + P8M6*P8_M6 + P9M6*P9_M6 + P10M6*P10_M6 + P11M6*P11_M6 +  P12M6*P2_M6 + P13M6*P13_M6 + P14M6*P14_M6 + P15M6*P15_M6 + P16M6*P16_M6 , "SM6")
model += (M[6] == P1M7*P1_M7 + P2M7*P2_M7 + P3M7*P3_M7 + P4M7*P4_M7 + P5M7*P5_M7 + P6M7*P6_M7 + P7M7*P7_M7 + P8M7*P8_M7 + P9M7*P9_M7 + P10M7*P10_M7 + P11M7*P11_M7 +  P12M7*P2_M7 + P13M7*P13_M7 + P14M7*P14_M7 + P15M7*P15_M7 + P16M7*P16_M7 , "SM7")

model += (X[0] == P1X1*P1_X1 + P2X1*P2_X1 + P3X1*P3_X1 + P4X1*P4_X1 + P5X1*P5_X1 + P6X1*P6_X1 + P7X1*P7_X1 + P8X1*P8_X1 + P9X1*P9_X1 + P10X1*P10_X1 + P11X1*P11_X1 +  P12X1*P2_X1 + P13X1*P13_X1 + P14X1*P14_X1 + P15X1*P15_X1 + P16X1*P16_X1 , "SX1")
model += (X[1] == P1X2*P1_X2 + P2X2*P2_X2 + P3X2*P3_X2 + P4X2*P4_X2 + P5X2*P5_X2 + P6X2*P6_X2 + P7X2*P7_X2 + P8X2*P8_X2 + P9X2*P9_X2 + P10X2*P10_X2 + P11X2*P11_X2 +  P12X2*P2_X2 + P13X2*P13_X2 + P14X2*P14_X2 + P15X2*P15_X2 + P16X2*P16_X2 , "SX2")
model += (X[2] == P1X3*P1_X3 + P2X3*P2_X3 + P3X3*P3_X3 + P4X3*P4_X3 + P5X3*P5_X3 + P6X3*P6_X3 + P7X3*P7_X3 + P8X3*P8_X3 + P9X3*P9_X3 + P10X3*P10_X3 + P11X3*P11_X3 +  P12X3*P2_X3 + P13X3*P13_X3 + P14X3*P14_X3 + P15X3*P15_X3 + P16X3*P16_X3 , "SX3")
model += (X[3] == P1X4*P1_X4 + P2X4*P2_X4 + P3X4*P3_X4 + P4X4*P4_X4 + P5X4*P5_X4 + P6X4*P6_X4 + P7X4*P7_X4 + P8X4*P8_X4 + P9X4*P9_X4 + P10X4*P10_X4 + P11X4*P11_X4 +  P12X4*P2_X4 + P13X4*P13_X4 + P14X4*P14_X4 + P15X4*P15_X4 + P16X4*P16_X4 , "SX4")
model += (X[4] == P1X5*P1_X5 + P2X5*P2_X5 + P3X5*P3_X5 + P4X5*P4_X5 + P5X5*P5_X5 + P6X5*P6_X5 + P7X5*P7_X5 + P8X5*P8_X5 + P9X5*P9_X5 + P10X5*P10_X5 + P11X5*P11_X5 +  P12X5*P2_X5 + P13X5*P13_X5 + P14X5*P14_X5 + P15X5*P15_X5 + P16X5*P16_X5 , "SX5")
model += (X[5] == P1X6*P1_X6 + P2X6*P2_X6 + P3X6*P3_X6 + P4X6*P4_X6 + P5X6*P5_X6 + P6X6*P6_X6 + P7X6*P7_X6 + P8X6*P8_X6 + P9X6*P9_X6 + P10X6*P10_X6 + P11X6*P11_X6 +  P12X6*P2_X6 + P13X6*P13_X6 + P14X6*P14_X6 + P15X6*P15_X6 + P16X6*P16_X6 , "SX6")
model += (X[6] == P1X7*P1_X7 + P2X7*P2_X7 + P3X7*P3_X7 + P4X7*P4_X7 + P5X7*P5_X7 + P6X7*P6_X7 + P7X7*P7_X7 + P8X7*P8_X7 + P9X7*P9_X7 + P10X7*P10_X7 + P11X7*P11_X7 +  P12X7*P2_X7 + P13X7*P13_X7 + P14X7*P14_X7 + P15X7*P15_X7 + P16X7*P16_X7 , "SX7")

model += (J[0] == P1J1*P1_J1 + P2J1*P2_J1 + P3J1*P3_J1 + P4J1*P4_J1 + P5J1*P5_J1 + P6J1*P6_J1 + P7J1*P7_J1 + P8J1*P8_J1 + P9J1*P9_J1 + P10J1*P10_J1 + P11J1*P11_J1 +  P12J1*P2_J1 + P13J1*P13_J1 + P14J1*P14_J1 + P15J1*P15_J1 + P16J1*P16_J1 , "SJ1")
model += (J[1] == P1J2*P1_J2 + P2J2*P2_J2 + P3J2*P3_J2 + P4J2*P4_J2 + P5J2*P5_J2 + P6J2*P6_J2 + P7J2*P7_J2 + P8J2*P8_J2 + P9J2*P9_J2 + P10J2*P10_J2 + P11J2*P11_J2 +  P12J2*P2_J2 + P13J2*P13_J2 + P14J2*P14_J2 + P15J2*P15_J2 + P16J2*P16_J2 , "SJ2")
model += (J[2] == P1J3*P1_J3 + P2J3*P2_J3 + P3J3*P3_J3 + P4J3*P4_J3 + P5J3*P5_J3 + P6J3*P6_J3 + P7J3*P7_J3 + P8J3*P8_J3 + P9J3*P9_J3 + P10J3*P10_J3 + P11J3*P11_J3 +  P12J3*P2_J3 + P13J3*P13_J3 + P14J3*P14_J3 + P15J3*P15_J3 + P16J3*P16_J3 , "SJ3")
model += (J[3] == P1J4*P1_J4 + P2J4*P2_J4 + P3J4*P3_J4 + P4J4*P4_J4 + P5J4*P5_J4 + P6J4*P6_J4 + P7J4*P7_J4 + P8J4*P8_J4 + P9J4*P9_J4 + P10J4*P10_J4 + P11J4*P11_J4 +  P12J4*P2_J4 + P13J4*P13_J4 + P14J4*P14_J4 + P15J4*P15_J4 + P16J4*P16_J4 , "SJ4")
model += (J[4] == P1J5*P1_J5 + P2J5*P2_J5 + P3J5*P3_J5 + P4J5*P4_J5 + P5J5*P5_J5 + P6J5*P6_J5 + P7J5*P7_J5 + P8J5*P8_J5 + P9J5*P9_J5 + P10J5*P10_J5 + P11J5*P11_J5 +  P12J5*P2_J5 + P13J5*P13_J5 + P14J5*P14_J5 + P15J5*P15_J5 + P16J5*P16_J5 , "SJ5")
model += (J[5] == P1J6*P1_J6 + P2J6*P2_J6 + P3J6*P3_J6 + P4J6*P4_J6 + P5J6*P5_J6 + P6J6*P6_J6 + P7J6*P7_J6 + P8J6*P8_J6 + P9J6*P9_J6 + P10J6*P10_J6 + P11J6*P11_J6 +  P12J6*P2_J6 + P13J6*P13_J6 + P14J6*P14_J6 + P15J6*P15_J6 + P16J6*P16_J6 , "SJ6")
model += (J[6] == P1J7*P1_J7 + P2J7*P2_J7 + P3J7*P3_J7 + P4J7*P4_J7 + P5J7*P5_J7 + P6J7*P6_J7 + P7J7*P7_J7 + P8J7*P8_J7 + P9J7*P9_J7 + P10J7*P10_J7 + P11J7*P11_J7 +  P12J7*P2_J7 + P13J7*P13_J7 + P14J7*P14_J7 + P15J7*P15_J7 + P16J7*P16_J7 , "SJ7")

model += (V[0] == P1V1*P1_V1 + P2V1*P2_V1 + P3V1*P3_V1 + P4V1*P4_V1 + P5V1*P5_V1 + P6V1*P6_V1 + P7V1*P7_V1 + P8V1*P8_V1 + P9V1*P9_V1 + P10V1*P10_V1 + P11V1*P11_V1 +  P12V1*P2_V1 + P13V1*P13_V1 + P14V1*P14_V1 + P15V1*P15_V1 + P16V1*P16_V1 , "SV1")
model += (V[1] == P1V2*P1_V2 + P2V2*P2_V2 + P3V2*P3_V2 + P4V2*P4_V2 + P5V2*P5_V2 + P6V2*P6_V2 + P7V2*P7_V2 + P8V2*P8_V2 + P9V2*P9_V2 + P10V2*P10_V2 + P11V2*P11_V2 +  P12V2*P2_V2 + P13V2*P13_V2 + P14V2*P14_V2 + P15V2*P15_V2 + P16V2*P16_V2 , "SV2")
model += (V[2] == P1V3*P1_V3 + P2V3*P2_V3 + P3V3*P3_V3 + P4V3*P4_V3 + P5V3*P5_V3 + P6V3*P6_V3 + P7V3*P7_V3 + P8V3*P8_V3 + P9V3*P9_V3 + P10V3*P10_V3 + P11V3*P11_V3 +  P12V3*P2_V3 + P13V3*P13_V3 + P14V3*P14_V3 + P15V3*P15_V3 + P16V3*P16_V3 , "SV3")
model += (V[3] == P1V4*P1_V4 + P2V4*P2_V4 + P3V4*P3_V4 + P4V4*P4_V4 + P5V4*P5_V4 + P6V4*P6_V4 + P7V4*P7_V4 + P8V4*P8_V4 + P9V4*P9_V4 + P10V4*P10_V4 + P11V4*P11_V4 +  P12V4*P2_V4 + P13V4*P13_V4 + P14V4*P14_V4 + P15V4*P15_V4 + P16V4*P16_V4 , "SV4")
model += (V[4] == P1V5*P1_V5 + P2V5*P2_V5 + P3V5*P3_V5 + P4V5*P4_V5 + P5V5*P5_V5 + P6V5*P6_V5 + P7V5*P7_V5 + P8V5*P8_V5 + P9V5*P9_V5 + P10V5*P10_V5 + P11V5*P11_V5 +  P12V5*P2_V5 + P13V5*P13_V5 + P14V5*P14_V5 + P15V5*P15_V5 + P16V5*P16_V5 , "SV5")
model += (V[5] == P1V6*P1_V6 + P2V6*P2_V6 + P3V6*P3_V6 + P4V6*P4_V6 + P5V6*P5_V6 + P6V6*P6_V6 + P7V6*P7_V6 + P8V6*P8_V6 + P9V6*P9_V6 + P10V6*P10_V6 + P11V6*P11_V6 +  P12V6*P2_V6 + P13V6*P13_V6 + P14V6*P14_V6 + P15V6*P15_V6 + P16V6*P16_V6 , "SV6")
model += (V[6] == P1V7*P1_V7 + P2V7*P2_V7 + P3V7*P3_V7 + P4V7*P4_V7 + P5V7*P5_V7 + P6V7*P6_V7 + P7V7*P7_V7 + P8V7*P8_V7 + P9V7*P9_V7 + P10V7*P10_V7 + P11V7*P11_V7 +  P12V7*P2_V7 + P13V7*P13_V7 + P14V7*P14_V7 + P15V7*P15_V7 + P16V7*P16_V7 , "SV7")

#2 primos por turno

model += ( 2 == P1_L1+P2_L1+P3_L1+P4_L1+P5_L1+P6_L1+P7_L1+P8_L1+P9_L1+P10_L1+P11_L1+P12_L1+P13_L1+P14_L1+P15_L1+P16_L1,"RL1")
model += ( 2 == P1_L2+P2_L2+P3_L2+P4_L2+P5_L2+P6_L2+P7_L2+P8_L2+P9_L2+P10_L2+P11_L2+P12_L2+P13_L2+P14_L2+P15_L2+P16_L2,"RL2")
model += ( 2 == P1_L3+P2_L3+P3_L3+P4_L3+P5_L3+P6_L3+P7_L3+P8_L3+P9_L3+P10_L3+P11_L3+P12_L3+P13_L3+P14_L3+P15_L3+P16_L3,"RL3")
model += ( 2 == P1_L4+P2_L4+P3_L4+P4_L4+P5_L4+P6_L4+P7_L4+P8_L4+P9_L4+P10_L4+P11_L4+P12_L4+P13_L4+P14_L4+P15_L4+P16_L4,"RL4")
model += ( 2 == P1_L5+P2_L5+P3_L5+P4_L5+P5_L5+P6_L5+P7_L5+P8_L5+P9_L5+P10_L5+P11_L5+P12_L5+P13_L5+P14_L5+P15_L5+P16_L5,"RL5")
model += ( 2 == P1_L6+P2_L6+P3_L6+P4_L6+P5_L6+P6_L6+P7_L6+P8_L6+P9_L6+P10_L6+P11_L6+P12_L6+P13_L6+P14_L6+P15_L6+P16_L6,"RL6")
model += ( 2 == P1_L7+P2_L7+P3_L7+P4_L7+P5_L7+P6_L7+P7_L7+P8_L7+P9_L7+P10_L7+P11_L7+P12_L7+P13_L7+P14_L7+P15_L7+P16_L7,"RL7")
model += ( 2 == P1_M1+P2_M1+P3_M1+P4_M1+P5_M1+P6_M1+P7_M1+P8_M1+P9_M1+P10_M1+P11_M1+P12_M1+P13_M1+P14_M1+P15_M1+P16_M1,"RM1")
model += ( 2 == P1_M2+P2_M2+P3_M2+P4_M2+P5_M2+P6_M2+P7_M2+P8_M2+P9_M2+P10_M2+P11_M2+P12_M2+P13_M2+P14_M2+P15_M2+P16_M2,"RM2")
model += ( 2 == P1_M3+P2_M3+P3_M3+P4_M3+P5_M3+P6_M3+P7_M3+P8_M3+P9_M3+P10_M3+P11_M3+P12_M3+P13_M3+P14_M3+P15_M3+P16_M3,"RM3")
model += ( 2 == P1_M4+P2_M4+P3_M4+P4_M4+P5_M4+P6_M4+P7_M4+P8_M4+P9_M4+P10_M4+P11_M4+P12_M4+P13_M4+P14_M4+P15_M4+P16_M4,"RM4")
model += ( 2 == P1_M5+P2_M5+P3_M5+P4_M5+P5_M5+P6_M5+P7_M5+P8_M5+P9_M5+P10_M5+P11_M5+P12_M5+P13_M5+P14_M5+P15_M5+P16_M5,"RM5")
model += ( 2 == P1_M6+P2_M6+P3_M6+P4_M6+P5_M6+P6_M6+P7_M6+P8_M6+P9_M6+P10_M6+P11_M6+P12_M6+P13_M6+P14_M6+P15_M6+P16_M6,"RM6")
model += ( 2 == P1_M7+P2_M7+P3_M7+P4_M7+P5_M7+P6_M7+P7_M7+P8_M7+P9_M7+P10_M7+P11_M7+P12_M7+P13_M7+P14_M7+P15_M7+P16_M7,"RM7")
model += ( 2 == P1_X1+P2_X1+P3_X1+P4_X1+P5_X1+P6_X1+P7_X1+P8_X1+P9_X1+P10_X1+P11_X1+P12_X1+P13_X1+P14_X1+P15_X1+P16_X1,"RX1")
model += ( 2 == P1_X2+P2_X2+P3_X2+P4_X2+P5_X2+P6_X2+P7_X2+P8_X2+P9_X2+P10_X2+P11_X2+P12_X2+P13_X2+P14_X2+P15_X2+P16_X2,"RX2")
model += ( 2 == P1_X3+P2_X3+P3_X3+P4_X3+P5_X3+P6_X3+P7_X3+P8_X3+P9_X3+P10_X3+P11_X3+P12_X3+P13_X3+P14_X3+P15_X3+P16_X3,"RX3")
model += ( 2 == P1_X4+P2_X4+P3_X4+P4_X4+P5_X4+P6_X4+P7_X4+P8_X4+P9_X4+P10_X4+P11_X4+P12_X4+P13_X4+P14_X4+P15_X4+P16_X4,"RX4")
model += ( 2 == P1_X5+P2_X5+P3_X5+P4_X5+P5_X5+P6_X5+P7_X5+P8_X5+P9_X5+P10_X5+P11_X5+P12_X5+P13_X5+P14_X5+P15_X5+P16_X5,"RX5")
model += ( 2 == P1_X6+P2_X6+P3_X6+P4_X6+P5_X6+P6_X6+P7_X6+P8_X6+P9_X6+P10_X6+P11_X6+P12_X6+P13_X6+P14_X6+P15_X6+P16_X6,"RX6")
model += ( 2 == P1_X7+P2_X7+P3_X7+P4_X7+P5_X7+P6_X7+P7_X7+P8_X7+P9_X7+P10_X7+P11_X7+P12_X7+P13_X7+P14_X7+P15_X7+P16_X7,"RX7")
model += ( 2 == P1_J1+P2_J1+P3_J1+P4_J1+P5_J1+P6_J1+P7_J1+P8_J1+P9_J1+P10_J1+P11_J1+P12_J1+P13_J1+P14_J1+P15_J1+P16_J1,"RJ1")
model += ( 2 == P1_J2+P2_J2+P3_J2+P4_J2+P5_J2+P6_J2+P7_J2+P8_J2+P9_J2+P10_J2+P11_J2+P12_J2+P13_J2+P14_J2+P15_J2+P16_J2,"RJ2")
model += ( 2 == P1_J3+P2_J3+P3_J3+P4_J3+P5_J3+P6_J3+P7_J3+P8_J3+P9_J3+P10_J3+P11_J3+P12_J3+P13_J3+P14_J3+P15_J3+P16_J3,"RJ3")
model += ( 2 == P1_J4+P2_J4+P3_J4+P4_J4+P5_J4+P6_J4+P7_J4+P8_J4+P9_J4+P10_J4+P11_J4+P12_J4+P13_J4+P14_J4+P15_J4+P16_J4,"RJ4")
model += ( 2 == P1_J5+P2_J5+P3_J5+P4_J5+P5_J5+P6_J5+P7_J5+P8_J5+P9_J5+P10_J5+P11_J5+P12_J5+P13_J5+P14_J5+P15_J5+P16_J5,"RJ5")
model += ( 2 == P1_J6+P2_J6+P3_J6+P4_J6+P5_J6+P6_J6+P7_J6+P8_J6+P9_J6+P10_J6+P11_J6+P12_J6+P13_J6+P14_J6+P15_J6+P16_J6,"RJ6")
model += ( 2 == P1_J7+P2_J7+P3_J7+P4_J7+P5_J7+P6_J7+P7_J7+P8_J7+P9_J7+P10_J7+P11_J7+P12_J7+P13_J7+P14_J7+P15_J7+P16_J7,"RJ7")
model += ( 2 == P1_V1+P2_V1+P3_V1+P4_V1+P5_V1+P6_V1+P7_V1+P8_V1+P9_V1+P10_V1+P11_V1+P12_V1+P13_V1+P14_V1+P15_V1+P16_V1,"RV1")
model += ( 2 == P1_V2+P2_V2+P3_V2+P4_V2+P5_V2+P6_V2+P7_V2+P8_V2+P9_V2+P10_V2+P11_V2+P12_V2+P13_V2+P14_V2+P15_V2+P16_V2,"RV2")
model += ( 2 == P1_V3+P2_V3+P3_V3+P4_V3+P5_V3+P6_V3+P7_V3+P8_V3+P9_V3+P10_V3+P11_V3+P12_V3+P13_V3+P14_V3+P15_V3+P16_V3,"RV3")
model += ( 2 == P1_V4+P2_V4+P3_V4+P4_V4+P5_V4+P6_V4+P7_V4+P8_V4+P9_V4+P10_V4+P11_V4+P12_V4+P13_V4+P14_V4+P15_V4+P16_V4,"RV4")
model += ( 2 == P1_V5+P2_V5+P3_V5+P4_V5+P5_V5+P6_V5+P7_V5+P8_V5+P9_V5+P10_V5+P11_V5+P12_V5+P13_V5+P14_V5+P15_V5+P16_V5,"RV5")
model += ( 2 == P1_V6+P2_V6+P3_V6+P4_V6+P5_V6+P6_V6+P7_V6+P8_V6+P9_V6+P10_V6+P11_V6+P12_V6+P13_V6+P14_V6+P15_V6+P16_V6,"RV6")
model += ( 2 == P1_V7+P2_V7+P3_V7+P4_V7+P5_V7+P6_V7+P7_V7+P8_V7+P9_V7+P10_V7+P11_V7+P12_V7+P13_V7+P14_V7+P15_V7+P16_V7,"RV7")

#maximo 4 o 5 horas por primo

model += ( P[1] == P1_L1+ P1_L2+ P1_L3+ P1_L4+ P1_L5+ P1_L6+ P1_L7+ P1_M1+ P1_M2+ P1_M3+ P1_M4+ P1_M5+ P1_M6+ P1_M7+ P1_X1+ P1_X2+ P1_X3+ P1_X4+ P1_X5+ P1_X6+ P1_X7+ P1_J1+ P1_J2+ P1_J3+ P1_J4+ P1_J5+ P1_J6+ P1_J7+ P1_V1+ P1_V2+ P1_V3+ P1_V4+ P1_V5+ P1_V6+ P1_V7, "RP1")
model += ( P[2] == P2_L1+ P2_L2+ P2_L3+ P2_L4+ P2_L5+ P2_L6+ P2_L7+ P2_M1+ P2_M2+ P2_M3+ P2_M4+ P2_M5+ P2_M6+ P2_M7+ P2_X1+ P2_X2+ P2_X3+ P2_X4+ P2_X5+ P2_X6+ P2_X7+ P2_J1+ P2_J2+ P2_J3+ P2_J4+ P2_J5+ P2_J6+ P2_J7+ P2_V1+ P2_V2+ P2_V3+ P2_V4+ P2_V5+ P2_V6+ P2_V7, "RP2")
model += ( P[3] == P3_L1+ P3_L2+ P3_L3+ P3_L4+ P3_L5+ P3_L6+ P3_L7+ P3_M1+ P3_M2+ P3_M3+ P3_M4+ P3_M5+ P3_M6+ P3_M7+ P3_X1+ P3_X2+ P3_X3+ P3_X4+ P3_X5+ P3_X6+ P3_X7+ P3_J1+ P3_J2+ P3_J3+ P3_J4+ P3_J5+ P3_J6+ P3_J7+ P3_V1+ P3_V2+ P3_V3+ P3_V4+ P3_V5+ P3_V6+ P3_V7, "RP3")
model += ( P[4] == P4_L1+ P4_L2+ P4_L3+ P4_L4+ P4_L5+ P4_L6+ P4_L7+ P4_M1+ P4_M2+ P4_M3+ P4_M4+ P4_M5+ P4_M6+ P4_M7+ P4_X1+ P4_X2+ P4_X3+ P4_X4+ P4_X5+ P4_X6+ P4_X7+ P4_J1+ P4_J2+ P4_J3+ P4_J4+ P4_J5+ P4_J6+ P4_J7+ P4_V1+ P4_V2+ P4_V3+ P4_V4+ P4_V5+ P4_V6+ P4_V7, "RP4")
model += ( P[5] == P5_L1+ P5_L2+ P5_L3+ P5_L4+ P5_L5+ P5_L6+ P5_L7+ P5_M1+ P5_M2+ P5_M3+ P5_M4+ P5_M5+ P5_M6+ P5_M7+ P5_X1+ P5_X2+ P5_X3+ P5_X4+ P5_X5+ P5_X6+ P5_X7+ P5_J1+ P5_J2+ P5_J3+ P5_J4+ P5_J5+ P5_J6+ P5_J7+ P5_V1+ P5_V2+ P5_V3+ P5_V4+ P5_V5+ P5_V6+ P5_V7, "RP5")
model += ( P[6] == P6_L1+ P6_L2+ P6_L3+ P6_L4+ P6_L5+ P6_L6+ P6_L7+ P6_M1+ P6_M2+ P6_M3+ P6_M4+ P6_M5+ P6_M6+ P6_M7+ P6_X1+ P6_X2+ P6_X3+ P6_X4+ P6_X5+ P6_X6+ P6_X7+ P6_J1+ P6_J2+ P6_J3+ P6_J4+ P6_J5+ P6_J6+ P6_J7+ P6_V1+ P6_V2+ P6_V3+ P6_V4+ P6_V5+ P6_V6+ P6_V7, "RP6")
model += ( P[7] == P7_L1+ P7_L2+ P7_L3+ P7_L4+ P7_L5+ P7_L6+ P7_L7+ P7_M1+ P7_M2+ P7_M3+ P7_M4+ P7_M5+ P7_M6+ P7_M7+ P7_X1+ P7_X2+ P7_X3+ P7_X4+ P7_X5+ P7_X6+ P7_X7+ P7_J1+ P7_J2+ P7_J3+ P7_J4+ P7_J5+ P7_J6+ P7_J7+ P7_V1+ P7_V2+ P7_V3+ P7_V4+ P7_V5+ P7_V6+ P7_V7, "RP7")
model += ( P[8] == P8_L1+ P8_L2+ P8_L3+ P8_L4+ P8_L5+ P8_L6+ P8_L7+ P8_M1+ P8_M2+ P8_M3+ P8_M4+ P8_M5+ P8_M6+ P8_M7+ P8_X1+ P8_X2+ P8_X3+ P8_X4+ P8_X5+ P8_X6+ P8_X7+ P8_J1+ P8_J2+ P8_J3+ P8_J4+ P8_J5+ P8_J6+ P8_J7+ P8_V1+ P8_V2+ P8_V3+ P8_V4+ P8_V5+ P8_V6+ P8_V7, "RP8")
model += ( P[9] == P9_L1+ P9_L2+ P9_L3+ P9_L4+ P9_L5+ P9_L6+ P9_L7+ P9_M1+ P9_M2+ P9_M3+ P9_M4+ P9_M5+ P9_M6+ P9_M7+ P9_X1+ P9_X2+ P9_X3+ P9_X4+ P9_X5+ P9_X6+ P9_X7+ P9_J1+ P9_J2+ P9_J3+ P9_J4+ P9_J5+ P9_J6+ P9_J7+ P9_V1+ P9_V2+ P9_V3+ P9_V4+ P9_V5+ P9_V6+ P9_V7, "RP9")
model += ( P[10] == P10_L1+ P10_L2+ P10_L3+ P10_L4+ P10_L5+ P10_L6+ P10_L7+ P10_M1+ P10_M2+ P10_M3+ P10_M4+ P10_M5+ P10_M6+ P10_M7+ P10_X1+ P10_X2+ P10_X3+ P10_X4+ P10_X5+ P10_X6+ P10_X7+ P10_J1+ P10_J2+ P10_J3+ P10_J4+ P10_J5+ P10_J6+ P10_J7+ P10_V1+ P10_V2+ P10_V3+ P10_V4+ P10_V5+ P10_V6+ P10_V7, "RP10")
model += ( P[11] == P11_L1+ P11_L2+ P11_L3+ P11_L4+ P11_L5+ P11_L6+ P11_L7+ P11_M1+ P11_M2+ P11_M3+ P11_M4+ P11_M5+ P11_M6+ P11_M7+ P11_X1+ P11_X2+ P11_X3+ P11_X4+ P11_X5+ P11_X6+ P11_X7+ P11_J1+ P11_J2+ P11_J3+ P11_J4+ P11_J5+ P11_J6+ P11_J7+ P11_V1+ P11_V2+ P11_V3+ P11_V4+ P11_V5+ P11_V6+ P11_V7, "RP11")
model += ( P[12] == P12_L1+ P12_L2+ P12_L3+ P12_L4+ P12_L5+ P12_L6+ P12_L7+ P12_M1+ P12_M2+ P12_M3+ P12_M4+ P12_M5+ P12_M6+ P12_M7+ P12_X1+ P12_X2+ P12_X3+ P12_X4+ P12_X5+ P12_X6+ P12_X7+ P12_J1+ P12_J2+ P12_J3+ P12_J4+ P12_J5+ P12_J6+ P12_J7+ P12_V1+ P12_V2+ P12_V3+ P12_V4+ P12_V5+ P12_V6+ P12_V7, "RP12")
model += ( P[13] == P13_L1+ P13_L2+ P13_L3+ P13_L4+ P13_L5+ P13_L6+ P13_L7+ P13_M1+ P13_M2+ P13_M3+ P13_M4+ P13_M5+ P13_M6+ P13_M7+ P13_X1+ P13_X2+ P13_X3+ P13_X4+ P13_X5+ P13_X6+ P13_X7+ P13_J1+ P13_J2+ P13_J3+ P13_J4+ P13_J5+ P13_J6+ P13_J7+ P13_V1+ P13_V2+ P13_V3+ P13_V4+ P13_V5+ P13_V6+ P13_V7, "RP13")
model += ( P[14] == P14_L1+ P14_L2+ P14_L3+ P14_L4+ P14_L5+ P14_L6+ P14_L7+ P14_M1+ P14_M2+ P14_M3+ P14_M4+ P14_M5+ P14_M6+ P14_M7+ P14_X1+ P14_X2+ P14_X3+ P14_X4+ P14_X5+ P14_X6+ P14_X7+ P14_J1+ P14_J2+ P14_J3+ P14_J4+ P14_J5+ P14_J6+ P14_J7+ P14_V1+ P14_V2+ P14_V3+ P14_V4+ P14_V5+ P14_V6+ P14_V7, "RP14")
model += ( P[15] == P15_L1+ P15_L2+ P15_L3+ P15_L4+ P15_L5+ P15_L6+ P15_L7+ P15_M1+ P15_M2+ P15_M3+ P15_M4+ P15_M5+ P15_M6+ P15_M7+ P15_X1+ P15_X2+ P15_X3+ P15_X4+ P15_X5+ P15_X6+ P15_X7+ P15_J1+ P15_J2+ P15_J3+ P15_J4+ P15_J5+ P15_J6+ P15_J7+ P15_V1+ P15_V2+ P15_V3+ P15_V4+ P15_V5+ P15_V6+ P15_V7, "RP15")
model += ( P[16] == P16_L1+ P16_L2+ P16_L3+ P16_L4+ P16_L5+ P16_L6+ P16_L7+ P16_M1+ P16_M2+ P16_M3+ P16_M4+ P16_M5+ P16_M6+ P16_M7+ P16_X1+ P16_X2+ P16_X3+ P16_X4+ P16_X5+ P16_X6+ P16_X7+ P16_J1+ P16_J2+ P16_J3+ P16_J4+ P16_J5+ P16_J6+ P16_J7+ P16_V1+ P16_V2+ P16_V3+ P16_V4+ P16_V5+ P16_V6+ P16_V7, "RP16")

#preferencia si quieren turnos juntos (PENDIENTE)

status = model.solve()

print(f"status: {model.status}, {LpStatus[model.status]}")
print(f"objective: {model.objective.value()}")

for var in model.variables():
    print(f"{var.name}: {var.value()}")

horario_final = [[[] for i in range (5)] for i in range (7)]

for var in model.variables():
    if re.match(r'P\d+_[LMXJV]\d+', var.name):
        if var.value() != 0:
            Bloque = var.name
            Aux = Bloque.split("_")
            Primo_de_turno = Aux[0][1:]
            Dia = [Aux[1][0],int(Aux[1][1])]
            Dia[0] = dias.index(Dia[0])
            Primo_de_turno = int(Primo_de_turno)-1
            nombre = Horarios_primos[Primo_de_turno][0]
            nombre = nombre.split(" ")
            nombre = nombre[0] + " " + nombre[2]
            horario_final[Dia[1]-1][Dia[0]].append(nombre)
        
workbook = openpyxl.load_workbook("Disponibilidad horaria primos SJ 2023-1.xlsx")

sheet = workbook.get_sheet_by_name("RESULTADOS")

matriz = horario_final

# Recorrer la matriz y escribir los datos en la hoja de cálculo
for row in range(7):
    for col in range(5):
         # Obtener el valor actual de la matriz
        valor_actual = matriz[row][col]
        
        row+=1
        col+=2

        if row >= 5:
            sheet.cell(row=row*2+2, column=col+1, value=valor_actual[0])
            sheet.cell(row=row*2+3, column=col+1, value=valor_actual[1])
        else:
            # Escribir el primer valor en la celda correspondiente
            sheet.cell(row=row*2+1, column=col+1, value=valor_actual[0])
            # Escribir el segundo valor en la celda siguiente
            sheet.cell(row=row*2+2, column=col+1, value=valor_actual[1])

        row-=1
        col-=2

# Guardar el archivo de Excel
workbook.save('matriz.xlsx')
